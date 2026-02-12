"""Breakdown export command: convert breakdown.json → breakdown.xlsx."""

import json
import click

from core.config import get_output_path


def run(proj: dict) -> None:
    """Read breakdown.json and write breakdown.xlsx."""
    project_name = proj["project"]
    click.secho(f"\n  Exporting breakdown for '{project_name}'", bold=True)

    # Load breakdown
    breakdown = _load_breakdown(proj)
    if not breakdown:
        return

    totals = _compute_totals(breakdown)
    _print_summary(breakdown, totals)

    # Write Excel
    xlsx_path = get_output_path(proj, "breakdown.xlsx")
    _write_excel(breakdown, totals, xlsx_path, project_name)

    click.secho(f"\n  ✓ Export complete", fg="green", bold=True)
    click.echo(f"    Excel: {xlsx_path}")


def _compute_totals(breakdown: dict) -> dict:
    """Compute effort totals by discipline."""
    totals = {"FE": 0, "BE": 0, "DevOps": 0, "Design": 0, "stories": 0}
    for epic in breakdown.get("epics", []):
        for feature in epic.get("features", []):
            for story in feature.get("stories", []):
                totals["FE"] += story.get("fe_days", 0)
                totals["BE"] += story.get("be_days", 0)
                totals["DevOps"] += story.get("devops_days", 0)
                totals["Design"] += story.get("design_days", 0)
                totals["stories"] += 1
    return totals


def _print_summary(breakdown: dict, totals: dict) -> None:
    """Print breakdown summary."""
    total_days = sum(totals.get(k, 0) for k in ["FE", "BE", "DevOps", "Design"])

    click.secho(f"\n  Summary:", bold=True)
    click.echo(f"    Stories: {totals['stories']}")
    click.echo(f"    Total days: {total_days}")

    click.secho(f"\n  Effort by discipline:", bold=True)
    for key in ["FE", "BE", "DevOps", "Design"]:
        days = totals.get(key, 0)
        click.echo(f"    {key:10s} {days:6.1f} days")

    click.secho(f"    {'TOTAL':10s} {total_days:6.1f} days", bold=True)

    click.secho(f"\n  By epic:", bold=True)
    for epic in breakdown.get("epics", []):
        epic_days = 0
        epic_stories = 0
        for feature in epic.get("features", []):
            for story in feature.get("stories", []):
                epic_days += sum(story.get(f"{k}_days", 0) for k in ["fe", "be", "devops", "design"])
                epic_stories += 1
        click.echo(f"    {epic.get('name', '?'):35s} {epic_stories:3d} stories, {epic_days:6.1f} days")


def _write_excel(breakdown: dict, totals: dict, path, project_name: str) -> None:
    """Write breakdown to Excel with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        click.secho("  ⚠ openpyxl not installed, skipping Excel. Run: pip install openpyxl", fg="yellow")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Breakdown"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Epic", "Feature", "User Story",
        "Acceptance Criteria", "FE (Days)", "BE (Days)",
        "DevOps (Days)", "Design (Days)", "Risks",
        "Comments", "Assumptions",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    row = 2
    for epic in breakdown.get("epics", []):
        epic_name = epic.get("name", "")
        for feature in epic.get("features", []):
            feat_name = feature.get("name", "")
            for story in feature.get("stories", []):
                values = [
                    epic_name,
                    feat_name,
                    story.get("title", ""),
                    story.get("acceptance_criteria", ""),
                    story.get("fe_days", 0),
                    story.get("be_days", 0),
                    story.get("devops_days", 0),
                    story.get("design_days", 0),
                    story.get("risks", ""),
                    story.get("comments", ""),
                    story.get("assumptions", ""),
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=3, value="TOTALS").font = Font(bold=True)
    for col, key in [(5, "FE"), (6, "BE"), (7, "DevOps"), (8, "Design")]:
        cell = ws.cell(row=row, column=col, value=totals.get(key, 0))
        cell.font = Font(bold=True)
        cell.border = thin_border

    total_days = sum(totals.get(k, 0) for k in ["FE", "BE", "DevOps", "Design"])
    ws.cell(row=row + 1, column=3, value="TOTAL DAYS").font = Font(bold=True)
    ws.cell(row=row + 1, column=5, value=total_days).font = Font(bold=True)

    # Column widths
    widths = [22, 22, 35, 40, 10, 10, 12, 12, 25, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(path)


def _load_breakdown(proj: dict) -> dict | None:
    """Load and validate breakdown JSON."""
    path = get_output_path(proj, "breakdown.json")
    if not path.exists():
        click.secho("  ✗ breakdown.json not found.", fg="red")
        click.echo("    Generate it in conversation first.")
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if "epics" not in data:
            click.secho("  ✗ Invalid breakdown.json format (missing 'epics' key)", fg="red")
            return None
        return data
    except json.JSONDecodeError as e:
        click.secho(f"  ✗ Failed to parse breakdown.json: {e}", fg="red")
        return None
